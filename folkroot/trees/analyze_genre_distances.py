"""
BSD 2-Clause License

Copyright (c) 2025, Hilda Romero-Velo
All rights reserved.

Redistribution and use in source and binary forms, with or without
modification, are permitted provided that the following conditions are met:

* Redistributions of source code must retain the above copyright notice, this
  list of conditions and the following disclaimer.

* Redistributions in binary form must reproduce the above copyright notice, this
  list of conditions and the following disclaimer in the documentation
  and/or other materials provided with the distribution.

THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
"""

"""
  Created by Hilda Romero-Velo on March 2025.
"""

"""
  Analyze inter-genre distances in a phylogenetic tree and generate a genre-level tree.
  Optimized version with parallelization for better performance.
"""

import multiprocessing
from functools import partial
from collections import defaultdict
import os
import sys
import argparse
import numpy as np
import pandas as pd
from trees_utils import set_all_seeds
from analysis_utils.genre_tree_builder import (
    build_genre_tree,
    calculate_genre_distances,
)
from analysis_utils.metrics_analysis import (
    calculate_genre_separation_ratio,
)
from analysis_utils.data_processing import (
    extract_level_feature,
    generate_distance_heatmap,
    find_tree_files,
)


def normalize_distance_matrix(distance_matrix):
    """
    Normalize distance matrix values to range [0-1].
    """
    if isinstance(distance_matrix, pd.DataFrame):
        values = distance_matrix.values
    else:
        values = distance_matrix

    min_val = np.min(values)
    max_val = np.max(values)

    if max_val > min_val:
        normalized = (values - min_val) / (max_val - min_val)
    else:
        normalized = np.zeros_like(values)

    if isinstance(distance_matrix, pd.DataFrame):
        return pd.DataFrame(
            normalized, index=distance_matrix.index, columns=distance_matrix.columns
        )
    else:
        return normalized


def save_genre_excel(
    distance_matrix,
    normalized_matrix,
    genre_counts,
    tree_file,
    feature,
    level,
    genre_separation_ratio,
    excel_path,
):
    """Save genre analysis results to an Excel file with multiple sheets."""
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Write distance matrices with 3 decimal rounding
        distance_matrix.round(3).to_excel(writer, sheet_name="Distances_Original")
        normalized_matrix.round(3).to_excel(writer, sheet_name="Distances_Normalized")

        # Add genre counts
        counts_df = pd.DataFrame(list(genre_counts.items()), columns=["Genre", "Count"])
        counts_df.to_excel(writer, sheet_name="Genre Counts", index=False)

        # Add metadata
        metadata = {
            "Tree File": tree_file,
            "Feature": feature,
            "Level": level,
            "Total Genres": len(genre_counts),
        }
        meta_df = pd.DataFrame(list(metadata.items()), columns=["Property", "Value"])
        meta_df.to_excel(writer, sheet_name="Metadata", index=False)

        # Add Genre Separation Ratio by genre
        if genre_separation_ratio:
            gsr_df = pd.DataFrame(
                list(genre_separation_ratio.items()),
                columns=["Genre", "Genre Separation Ratio"],
            )
            gsr_df.sort_values("Genre Separation Ratio", ascending=False, inplace=True)
            gsr_df.to_excel(writer, sheet_name="Genre Separation Ratio", index=False)

        # Add normalization information
        norm_info = {
            "Original min": distance_matrix.values.min(),
            "Original max": distance_matrix.values.max(),
            "Normalized min": normalized_matrix.values.min(),
            "Normalized max": normalized_matrix.values.max(),
            "Note": "Values have been normalized to range [0-1] where 0 = most similar, 1 = most different",
        }
        norm_df = pd.DataFrame(list(norm_info.items()), columns=["Metric", "Value"])
        norm_df.to_excel(writer, sheet_name="Normalization Info", index=False)

    print(f"Genre distance matrix saved to {excel_path}")


def format_level_for_display(level):
    """Format level name for display in titles and charts."""
    if level == "note":
        level = "global"

    if level == "shared_segments":
        level = "Shared Phrases"

    if level == "structure":
        level = "form"

    if "s25_ss75" in level:
        level = level.replace("s25_ss75", "F25_S75")
    elif "s50_ss50" in level:
        level = level.replace("s50_ss50", "F50_S50")
    elif "s75_ss25" in level:
        level = level.replace("s75_ss25", "F75_S25")

    level = level.replace("_", " ")

    if level and len(level) > 0:
        level = level[0].upper() + level[1:]

    return level


def format_feature_for_display(feature):
    """Format feature name for display in titles and charts."""
    parts = feature.split("_")
    parts = [part.capitalize() for part in parts]
    return "-".join(parts)


def process_tree_analysis(tree_file, db_path, seed=42):
    """Core function to process a tree and generate all artifacts."""
    tree_name = os.path.basename(tree_file)
    tree_base = os.path.splitext(tree_name)[0]

    # Extract level and feature
    level, feature = extract_level_feature(tree_base)

    print(f"Analyzing {tree_name} (level={level}, feature={feature})...")

    # Calculate genre distances
    distance_matrix, genre_counts, metrics_data = calculate_genre_distances(
        tree_file, db_path
    )
    if not isinstance(distance_matrix, pd.DataFrame):
        genres = list(genre_counts.keys())
        distance_matrix = pd.DataFrame(distance_matrix, index=genres, columns=genres)
        print("Warning: Converted distance matrix to DataFrame for processing")

    normalized_matrix = normalize_distance_matrix(distance_matrix)

    if isinstance(normalized_matrix, np.ndarray):
        genre_names = list(genre_counts.keys())
        normalized_matrix = pd.DataFrame(
            normalized_matrix, index=genre_names, columns=genre_names
        )

    # Calculate genre separation ratio
    genre_separation_ratio = calculate_genre_separation_ratio(tree_file, db_path)

    # Create output directory
    output_dir = os.path.join(os.path.dirname(tree_file), "genre_analysis")
    os.makedirs(output_dir, exist_ok=True)

    # Save genre data to Excel
    excel_path = os.path.join(output_dir, f"genre_distances_{level}_{feature}.xlsx")
    save_genre_excel(
        distance_matrix,
        normalized_matrix,
        genre_counts,
        tree_file,
        feature,
        level,
        genre_separation_ratio,
        excel_path,
    )

    # Generate heatmap
    heatmap_path = os.path.join(output_dir, f"heatmap_{level}_{feature}.png")
    heatmap_title = f"{format_feature_for_display(feature)} Genre Distance - {format_level_for_display(level)} Similarity"
    generate_distance_heatmap(
        normalized_matrix, heatmap_path, title=heatmap_title, db_path=db_path
    )
    print(f"Heatmap visualization saved to {heatmap_path}")

    # Generate genre-level tree
    tree_output = os.path.join(output_dir, f"genre_tree_{level}_{feature}.nexus")
    build_genre_tree(distance_matrix, tree_output, random_seed=seed)
    print(f"Genre-level tree saved to {tree_output}")

    # Prepare data for comparison
    summary = {
        "Tree": tree_base,
        "Level": level,
        "Feature": feature,
        "Genres": len(genre_counts),
    }

    # Prepare genre data for comparative analysis
    genre_data = {}
    key_gsr = f"{level}_{feature}_gsr"

    all_genres = set(genre_separation_ratio.keys())

    for genre in all_genres:
        genre_data[genre] = {}
        if genre in genre_separation_ratio:
            genre_data[genre][key_gsr] = genre_separation_ratio[genre]

    return (
        summary,
        genre_data,
        level,
        feature,
        distance_matrix,
        normalized_matrix,
        genre_counts,
    )


def process_tree(tree_file, db_path):
    """Process a single tree in parallel for multi-tree analysis."""
    try:
        summary, genre_data, *_ = process_tree_analysis(tree_file, db_path)
        return summary, genre_data
    except Exception as e:
        print(f"Error analyzing tree {tree_file}: {e}")
        return None, None


def analyze_multiple_trees(tree_files, db_path, num_workers=None):
    """Analyze multiple trees in parallel and generate comparative metrics."""
    # Determine number of worker processes
    if num_workers is None:
        num_workers = max(1, multiprocessing.cpu_count() - 1)  # Leave one core free

    print(f"Found {len(tree_files)} trees to analyze")
    print(f"Processing trees in parallel using {num_workers} workers...")

    # Set up parallel processing
    process_tree_with_db = partial(process_tree, db_path=db_path)

    all_results = []
    all_genre_scores = defaultdict(dict)

    # Process trees in parallel
    with multiprocessing.Pool(processes=num_workers) as pool:
        results = pool.map(process_tree_with_db, tree_files)

    # Consolidate results
    for result, genre_data in results:
        if result:
            all_results.append(result)
            # Update genre scores dictionary
            for genre, scores in genre_data.items():
                all_genre_scores[genre].update(scores)

    return all_results, all_genre_scores


def analyze_single_tree(tree_file, db_path, seed=42):
    """Analize a single tree and generate all artifacts."""
    _, _, level, feature, distance_matrix, normalized_matrix, genre_counts = (
        process_tree_analysis(tree_file, db_path, seed)
    )
    return level, feature, distance_matrix, normalized_matrix, genre_counts


def save_comparison_results(all_results, all_genre_scores, output_dir):
    """
    Save comparative tables of Genre Separation Ratios by genre and configuration.

    Args:
        all_results (list): List of dictionaries with results by tree (used for future extensions)
        all_genre_scores (dict): Dictionary of scores by genre
        output_dir (str): Directory where to save results

    Returns:
        str: Path to generated Excel file
    """
    comparison_path = os.path.join(output_dir, "genre_metrics_comparison.xlsx")
    print(f"Saving genre comparison metrics to {comparison_path}...")

    # Define categories and types for systematic column ordering
    categories = {
        "base": ["note", "structure", "shared_segments"],
        "combined_weights": ["s25_ss75", "s50_ss50", "s75_ss25"],
    }

    feature_types = [
        "diatonic",
        "chromatic",
        "rhythmic",
        "diatonic_rhythmic",
        "chromatic_rhythmic",
    ]

    # Generate column order systematically
    base_column_order = []

    # First add base categories in order
    for category in categories["base"]:
        for feature in feature_types:
            base_column_order.append(f"{category}_{feature}")

    # Then add combined features
    for weight in categories["combined_weights"]:
        for feature in feature_types:
            base_column_order.append(f"combined_{weight}_{feature}")

    # Define metrics with their properties
    metrics_data = {
        "gsr": {
            "rows": [],
            "sheet": "Genre_Separation_Ratio",
            "note": "For Genre Separation Ratio, HIGHER values indicate better separation between genres",
        },
    }

    # Process each genre's data for all metrics
    for genre, scores in all_genre_scores.items():
        for metric_key, metric_info in metrics_data.items():
            row = {"Genre": genre}

            # Add metrics for each column if they exist
            for base_col in base_column_order:
                col_key = f"{base_col}_{metric_key}"
                if col_key in scores:
                    row[base_col] = (
                        round(scores[col_key], 4)
                        if isinstance(scores[col_key], (float, np.floating))
                        else scores[col_key]
                    )

            # Only add rows with actual data
            if len(row) > 1:
                metric_info["rows"].append(row)

    # Create DataFrames and prepare for Excel
    for metric_key, metric_info in metrics_data.items():
        if metric_info["rows"]:
            df = pd.DataFrame(metric_info["rows"])

            # Keep consistent column order where available
            cols = ["Genre"] + [col for col in base_column_order if col in df.columns]
            df = df[cols]
            df.sort_values("Genre", inplace=True)

            # Store DataFrame for writing to Excel
            metrics_data[metric_key]["dataframe"] = df

    # Check if we have data to save
    if not any("dataframe" in metric_info for metric_info in metrics_data.values()):
        print("No comparison data available to save")
        return comparison_path

    # Write to Excel with formatting
    with pd.ExcelWriter(comparison_path, engine="openpyxl") as writer:
        from openpyxl.styles import Font

        # Write each metric to its own sheet
        for metric_key, metric_info in metrics_data.items():
            if "dataframe" in metric_info:
                df = metric_info["dataframe"]
                df.to_excel(writer, sheet_name=metric_info["sheet"], index=False)

                # Apply formatting
                workbook = writer.book
                worksheet = workbook[metric_info["sheet"]]

                # Bold headers
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=1, column=col).font = Font(bold=True)

                # Add explanatory note
                note_row = len(df) + 3
                worksheet.cell(
                    row=note_row, column=1, value=f"Note: {metric_info['note']}"
                )
                worksheet.cell(row=note_row, column=1).font = Font(italic=True)

    print(f"Comparison of genre metrics saved to {comparison_path}")
    return comparison_path


if __name__ == "__main__":
    """Main script for analyzing genre distances in phylogenetic trees."""
    parser = argparse.ArgumentParser(
        description="Analyze genre distances in phylogenetic trees. Generate genre-level trees and comparison reports."
    )

    # Create mutually exclusive group for the analysis mode
    analysis_group = parser.add_mutually_exclusive_group(required=True)
    analysis_group.add_argument(
        "--tree-file",
        type=str,
        help="Path to a single NEXUS tree file. Generate genre tree, heatmap, and metrics.",
    )
    analysis_group.add_argument(
        "--directory",
        type=str,
        help="Directory containing tree files. Analyze all trees and create a comparison report.",
    )

    parser.add_argument(
        "--seed", type=int, default=42, help="Random seed for reproducibility"
    )
    parser.add_argument(
        "--num-workers",
        type=int,
        default=None,
        help="Number of worker processes (default: number of CPU cores - 1)",
    )

    args = parser.parse_args()
    set_all_seeds(args.seed)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.abspath(os.path.join(script_dir, "../database/folkroot.db"))

    if args.directory:  # Multiple tree analysis
        tree_files = find_tree_files(args.directory)
        if not tree_files:
            print(f"No suitable tree files found in directory: {args.directory}")
            sys.exit(1)

        # Process trees in parallel
        all_results, all_genre_scores = analyze_multiple_trees(
            tree_files, db_path, args.num_workers
        )

        # Create output directory and save results
        output_dir = os.path.join(script_dir, "genre_analysis")
        os.makedirs(output_dir, exist_ok=True)

        # Save comparative results with all metrics
        save_comparison_results(all_results, all_genre_scores, output_dir)

        print("Multi-tree analysis and metrics comparison completed successfully.")
    else:  # Single tree analysis
        level, feature, distance_matrix, normalized_matrix, genre_counts = (
            analyze_single_tree(args.tree_file, db_path, args.seed)
        )

        print(f"Single tree analysis for {args.tree_file} completed successfully.")
